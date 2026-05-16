import openpyxl
from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Q
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import authenticate, login, logout
from .models import Sapi, Dokter, Pemeliharaan, Supervisi, Pengeluaran, Panen
from .forms import SapiForm, DokterForm, PemeliharaanForm, SupervisiForm, PengeluaranForm, PanenForm, LaporanForm


# ==========================================
# AUTH (LOGIN & LOGOUT)
# ==========================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('peternakan:dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('peternakan:dashboard')
        else:
            messages.error(request, 'Username atau Password salah!')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.success(request, 'Anda telah berhasil logout.')
    return redirect('/login/')

# ==========================================
# DASHBOARD
# ==========================================
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'peternakan/dashboard.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bulan_ini = timezone.now().month
        tahun_ini = timezone.now().year
        
        context['total_sapi'] = Sapi.objects.filter(status='Hidup/Sehat').count()
        context['sapi_sakit'] = Sapi.objects.filter(status='Sakit').count()
        
        panen = Panen.objects.filter(tanggal__month=bulan_ini, tanggal__year=tahun_ini)
        context['pendapatan'] = sum([p.total_pendapatan for p in panen])
        
        pengeluaran = Pengeluaran.objects.filter(tanggal__month=bulan_ini, tanggal__year=tahun_ini)
        pakan = Pemeliharaan.objects.filter(tanggal__month=bulan_ini, tanggal__year=tahun_ini)
        medis = Supervisi.objects.filter(tanggal__month=bulan_ini, tanggal__year=tahun_ini)
        
        context['pengeluaran'] = (pengeluaran.aggregate(t=Sum('jumlah'))['t'] or 0) + \
                                 (pakan.aggregate(t=Sum('biaya_pakan'))['t'] or 0) + \
                                 (medis.aggregate(t=Sum('biaya_konsultasi'))['t'] or 0)
        context['laba'] = context['pendapatan'] - context['pengeluaran']
        return context

# ==========================================
# MODUL SAPI
# ==========================================
class SapiListView(LoginRequiredMixin, ListView):
    model = Sapi
    template_name = 'peternakan/sapi/list.html'
    context_object_name = 'data'
    paginate_by = 10
    login_url = '/login/'

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        if search: qs = qs.filter(id_sapi__icontains=search) | qs.filter(nama_sapi__icontains=search)
        if status: qs = qs.filter(status=status)
        return qs

class SapiCreateView(LoginRequiredMixin, CreateView):
    model = Sapi
    form_class = SapiForm
    template_name = 'peternakan/sapi/form.html'
    login_url = '/login/'
    def get_success_url(self):
        messages.success(self.request, 'Data sapi berhasil ditambahkan.')
        return reverse_lazy('peternakan:sapi_list')

class SapiUpdateView(LoginRequiredMixin, UpdateView):
    model = Sapi
    form_class = SapiForm
    template_name = 'peternakan/sapi/form.html'
    login_url = '/login/'
    def get_success_url(self):
        messages.success(self.request, 'Data sapi berhasil diupdate.')
        return reverse_lazy('peternakan:sapi_list')

class SapiDetailView(LoginRequiredMixin, DetailView):
    model = Sapi
    template_name = 'peternakan/sapi/detail.html'
    context_object_name = 'sapi'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sapi = self.get_object()
        context['pakan'] = sapi.riwayat_pakan.all().order_by('-tanggal')[:5]
        context['medis'] = sapi.riwayat_medis.all().order_by('-tanggal')[:5]
        context['total_pakan'] = sapi.riwayat_pakan.aggregate(t=Sum('biaya_pakan'))['t'] or 0
        context['total_medis'] = sapi.riwayat_medis.aggregate(t=Sum('biaya_konsultasi'))['t'] or 0
        return context

class SapiDeleteView(LoginRequiredMixin, DeleteView):
    model = Sapi
    template_name = 'peternakan/sapi/confirm_delete.html'
    success_url = reverse_lazy('peternakan:sapi_list')
    login_url = '/login/'
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Data sapi dihapus.')
        return super().delete(request, *args, **kwargs)

# ==========================================
# MODUL DOKTER
# ==========================================
class DokterListView(LoginRequiredMixin, ListView):
    model = Dokter
    template_name = 'peternakan/dokter/list.html'
    context_object_name = 'data'
    login_url = '/login/'

class DokterCreateView(LoginRequiredMixin, CreateView):
    model = Dokter
    form_class = DokterForm
    template_name = 'peternakan/dokter/form.html'
    success_url = reverse_lazy('peternakan:dokter_list')
    login_url = '/login/'

class DokterUpdateView(LoginRequiredMixin, UpdateView):
    model = Dokter
    form_class = DokterForm
    template_name = 'peternakan/dokter/form.html'
    success_url = reverse_lazy('peternakan:dokter_list')
    login_url = '/login/'

class DokterDeleteView(LoginRequiredMixin, DeleteView):
    model = Dokter
    template_name = 'peternakan/dokter/confirm_delete.html'
    success_url = reverse_lazy('peternakan:dokter_list')
    login_url = '/login/'

# ==========================================
# MODUL PEMELIHARAAN (SEARCH & SORT)
# ==========================================
class PemeliharaanListView(LoginRequiredMixin, ListView):
    model = Pemeliharaan
    template_name = 'peternakan/pemeliharaan/list.html'
    context_object_name = 'data'
    paginate_by = 10
    login_url = '/login/'

    def get_queryset(self):
        qs = Pemeliharaan.objects.select_related('sapi')
        search = self.request.GET.get('search')
        sort_by = self.request.GET.get('sort_by')

        if search:
            qs = qs.filter(
                Q(sapi__id_sapi__icontains=search) | 
                Q(sapi__nama_sapi__icontains=search) | 
                Q(jenis_pakan__icontains=search)
            )

        if sort_by == 'sapi_az': qs = qs.order_by('sapi__nama_sapi')
        elif sort_by == 'sapi_za': qs = qs.order_by('-sapi__nama_sapi')
        elif sort_by == 'tgl_lama': qs = qs.order_by('tanggal')
        else: qs = qs.order_by('-tanggal')
        return qs

class PemeliharaanCreateView(LoginRequiredMixin, CreateView):
    model = Pemeliharaan
    form_class = PemeliharaanForm
    template_name = 'peternakan/pemeliharaan/form.html'
    success_url = reverse_lazy('peternakan:pemeliharaan_list')
    login_url = '/login/'

class PemeliharaanUpdateView(LoginRequiredMixin, UpdateView):
    model = Pemeliharaan
    form_class = PemeliharaanForm
    template_name = 'peternakan/pemeliharaan/form.html'
    success_url = reverse_lazy('peternakan:pemeliharaan_list')
    login_url = '/login/'

class PemeliharaanDeleteView(LoginRequiredMixin, DeleteView):
    model = Pemeliharaan
    template_name = 'peternakan/pemeliharaan/confirm_delete.html'
    success_url = reverse_lazy('peternakan:pemeliharaan_list')
    login_url = '/login/'

def export_pemeliharaan_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from django.db.models import Q
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Data_Pemeliharaan_Pakan.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pakan"

    # Styling Header
    header_fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True)
    
    columns = ['Tanggal', 'ID Sapi', 'Nama Sapi', 'Jenis Pakan', 'Jumlah (Kg)', 'Biaya (Rp)', 'Keterangan']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Filter Data
    search = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', 'tgl_baru')
    qs = Pemeliharaan.objects.select_related('sapi')
    if search: qs = qs.filter(Q(sapi__id_sapi__icontains=search) | Q(sapi__nama_sapi__icontains=search) | Q(jenis_pakan__icontains=search))
    if sort_by == 'sapi_az': qs = qs.order_by('sapi__nama_sapi')
    elif sort_by == 'sapi_za': qs = qs.order_by('-sapi__nama_sapi')
    elif sort_by == 'tgl_lama': qs = qs.order_by('tanggal')
    else: qs = qs.order_by('-tanggal')

    # Isi Data
    for row_num, p in enumerate(qs, 2):
        ws.cell(row=row_num, column=1, value=p.tanggal.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=2, value=p.sapi.id_sapi)
        ws.cell(row=row_num, column=3, value=p.sapi.nama_sapi)
        ws.cell(row=row_num, column=4, value=p.jenis_pakan)
        ws.cell(row=row_num, column=5, value=float(p.jumlah_pakan))
        ws.cell(row=row_num, column=6, value=float(p.biaya_pakan))
        ws.cell(row=row_num, column=7, value=p.keterangan if p.keterangan else "-")

    # Auto adjust lebar kolom
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

# ==========================================
# MODUL SUPERVISI (SEARCH, SORT, UPDATE FISIK)
# ==========================================
class SupervisiListView(LoginRequiredMixin, ListView):
    model = Supervisi
    template_name = 'peternakan/supervisi/list.html'
    context_object_name = 'data'
    paginate_by = 10
    login_url = '/login/'

    def get_queryset(self):
        qs = Supervisi.objects.select_related('sapi', 'dokter')
        search = self.request.GET.get('search')
        sort_by = self.request.GET.get('sort_by')

        if search:
            qs = qs.filter(
                Q(sapi__id_sapi__icontains=search) | 
                Q(sapi__nama_sapi__icontains=search) | 
                Q(diagnosa__icontains=search) |
                Q(dokter__nama__icontains=search)
            )

        if sort_by == 'sapi_az': qs = qs.order_by('sapi__nama_sapi')
        elif sort_by == 'sapi_za': qs = qs.order_by('-sapi__nama_sapi')
        elif sort_by == 'tgl_lama': qs = qs.order_by('tanggal')
        elif sort_by == 'dokter_az': qs = qs.order_by('dokter__nama')
        elif sort_by == 'diagnosa_az': qs = qs.order_by('diagnosa')
        else: qs = qs.order_by('-tanggal')
        return qs

class SupervisiCreateView(LoginRequiredMixin, CreateView):
    model = Supervisi
    form_class = SupervisiForm
    template_name = 'peternakan/supervisi/form.html'
    login_url = '/login/'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        sapi = form.instance.sapi
        sapi.status = form.instance.update_status_sapi
        if form.instance.berat_sapi_saat_ini: sapi.berat_badan = form.instance.berat_sapi_saat_ini
        if form.instance.update_tinggi_sapi: sapi.tinggi_sapi = form.instance.update_tinggi_sapi
        sapi.save(update_fields=['status', 'berat_badan', 'tinggi_sapi'])
        messages.success(self.request, f'Data {sapi.nama_sapi} diperbarui.')
        return response
    
    def get_success_url(self): return reverse_lazy('peternakan:supervisi_list')

# ==========================================
# MODUL PENGELUARAN
# ==========================================
class PengeluaranListView(LoginRequiredMixin, ListView):
    model = Pengeluaran
    template_name = 'peternakan/pengeluaran/list.html'
    context_object_name = 'data'
    paginate_by = 10
    login_url = '/login/'

    def get_queryset(self):
        qs = Pengeluaran.objects.all()
        search = self.request.GET.get('search')
        sort_by = self.request.GET.get('sort_by')

        if search:
            qs = qs.filter(Q(kategori__icontains=search) | Q(deskripsi__icontains=search))

        if sort_by == 'tgl_lama': qs = qs.order_by('tanggal')
        elif sort_by == 'kategori_az': qs = qs.order_by('kategori')
        elif sort_by == 'jumlah_termahal': qs = qs.order_by('-jumlah')
        elif sort_by == 'jumlah_termurah': qs = qs.order_by('jumlah')
        else: qs = qs.order_by('-tanggal')
        return qs

class PengeluaranCreateView(LoginRequiredMixin, CreateView):
    model = Pengeluaran
    form_class = PengeluaranForm
    template_name = 'peternakan/pengeluaran/form.html'
    success_url = reverse_lazy('peternakan:pengeluaran_list')
    login_url = '/login/'

class PengeluaranDeleteView(LoginRequiredMixin, DeleteView):
    model = Pengeluaran
    template_name = 'peternakan/pengeluaran/confirm_delete.html'
    success_url = reverse_lazy('peternakan:pengeluaran_list')
    login_url = '/login/'

# ==========================================
# MODUL PANEN (PASTIKAN BAGIAN INI ADA)
# ==========================================
class PanenListView(LoginRequiredMixin, ListView):
    model = Panen
    template_name = 'peternakan/panen/list.html'
    context_object_name = 'data'
    paginate_by = 10
    login_url = '/login/'

    def get_queryset(self):
        qs = Panen.objects.select_related('sapi')
        search = self.request.GET.get('search')
        sort_by = self.request.GET.get('sort_by')

        if search:
            qs = qs.filter(
                Q(jenis_produksi__icontains=search) | 
                Q(sapi__nama_sapi__icontains=search) | 
                Q(satuan__icontains=search)
            )

        if sort_by == 'tgl_lama': qs = qs.order_by('tanggal')
        elif sort_by == 'jenis_az': qs = qs.order_by('jenis_produksi')
        elif sort_by == 'pendapatan_terbesar': qs = qs.order_by('-jumlah')
        else: qs = qs.order_by('-tanggal')
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_pendapatan'] = sum([p.total_pendapatan for p in context['data']])
        return context

class PanenCreateView(LoginRequiredMixin, CreateView):
    model = Panen
    form_class = PanenForm
    template_name = 'peternakan/panen/form.html'
    login_url = '/login/'
    
    def form_valid(self, form):
        if form.cleaned_data['jenis_produksi'] == 'Penjualan Sapi Dewasa':
            sapi = form.cleaned_data.get('sapi')
            if sapi:
                sapi.status = 'Dijual'
                sapi.save(update_fields=['status'])
        response = super().form_valid(form)
        messages.success(self.request, 'Data panen/penjualan berhasil dicatat.')
        return response
    
    def get_success_url(self): return reverse_lazy('peternakan:panen_list')

class PanenDeleteView(LoginRequiredMixin, DeleteView):
    model = Panen
    template_name = 'peternakan/panen/confirm_delete.html'
    success_url = reverse_lazy('peternakan:panen_list')
    login_url = '/login/'

# ==========================================
# MODUL LAPORAN
# ==========================================
class LaporanView(LoginRequiredMixin, TemplateView):
    template_name = 'peternakan/laporan/index.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = LaporanForm()
        
        if 'tanggal_mulai' in self.request.GET and 'tanggal_selesai' in self.request.GET:
            form = LaporanForm(self.request.GET)
            if form.is_valid():
                mulai = form.cleaned_data['tanggal_mulai']
                selesai = form.cleaned_data['tanggal_selesai']
                
                panens = Panen.objects.filter(tanggal__range=(mulai, selesai))
                pengeluarans = Pengeluaran.objects.filter(tanggal__range=(mulai, selesai))
                pakans = Pemeliharaan.objects.filter(tanggal__range=(mulai, selesai))
                medis = Supervisi.objects.filter(tanggal__range=(mulai, selesai))
                
                total_pendapatan = sum([p.total_pendapatan for p in panens])
                total_pengeluaran = (pengeluarans.aggregate(t=Sum('jumlah'))['t'] or 0) + \
                                    (pakans.aggregate(t=Sum('biaya_pakan'))['t'] or 0) + \
                                    (medis.aggregate(t=Sum('biaya_konsultasi'))['t'] or 0)
                
                context['active'] = True
                context['pendapatan'] = total_pendapatan
                context['pengeluaran'] = total_pengeluaran
                context['laba'] = total_pendapatan - total_pengeluaran
                context['list_pendapatan'] = panens
                context['list_pengeluaran'] = pengeluarans
        return context